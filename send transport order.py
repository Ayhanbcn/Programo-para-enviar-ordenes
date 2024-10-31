import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
import os
import datetime
from datetime import datetime
import keyring
import re

trh = datetime.now().strftime("%Y-%m-%d")

# Check if the Excel file exists, if not, create it

excel_file = "Lista de envios.xlsx"

if not os.path.exists(excel_file):
    workbook = openpyxl.Workbook()

    sheet1 = workbook.active
    sheet1.title = "Envios"
    sheet1.append(["Fecha", "Cantidad de Palet", "Lugar de Recogida", "Lugar de Entrega", "Transporte", "Numeros de Palets", "Numero de Factura Transporte"])

    sheet2 = workbook.create_sheet("Clientes")
    sheet2.append(["Nombre", "Telefono", "Direccion"])

    sheet3 = workbook.create_sheet("Almacenes")
    sheet3.append(["Nombre de Nave", "Telefono", "Direccion"])

    sheet4 = workbook.create_sheet("Transportistas")
    sheet4.append(["Nombre de Transportiste", "Emails"])

    workbook.save(excel_file)

# Save email and password
def save_credentials():
    global email, password
    email = from_address_entry.get()
    password = password_entry.get()
    keyring.set_password("mail_app", "email", email)
    keyring.set_password("mail_app", "password", password)
    from_address_entry.delete(0, tk.END)
    password_entry.delete(0, tk.END)
    messagebox.showinfo("Éxito", "¡Información guardada!")

# Load email and password
def load_credentials():
    global email, password
    email = keyring.get_password("mail_app", "email")
    password = keyring.get_password("mail_app", "password")

    if email and password:
        from_address_entry.insert(0, email)
        password_entry.insert(0, password)

# Greet by time zone
def great():
    current_hour = datetime.now().hour

    if current_hour < 12:
        greeting = "¡Buenos días!"
    elif 12 <= current_hour < 18:
        greeting = "¡Buenas tardes!"
    else:
        greeting = "¡Buenas noches!"

    return greeting

#################################################
# Function that updates the email content
def update_email_content():
    greeting = great()
    subject1 = f"Recogida {palet.get()} palet en {pick_up_place.get()} para {place_to_delivery.get()}"
    content = (
    f"{greeting}\n\n"
    f"Por favor podeis recoger {palet.get()} palet en {pick_up_place.get()} para {place_to_delivery.get()} para {when.get()}\n\n"
    f"{palet_content_textbox.get(1.0, tk.END)}\n\n"
    f"Direccion de recogida:\n{pick_up_place.get()}\n{pick_up_address.get()}\n{pick_up_tel.get()}\n\n"
    f"Direccion de entrega:\n{place_to_delivery.get()}\n{delivery_address.get()}\n{delivery_tel.get()}\n\n"
    f"Adjunto los documentos,\n\n"
    f"Saludos,"
)
    email_content_textbox.delete(1.0, tk.END)
    email_content_textbox.insert(tk.END, content)
    email_subject_textbox.delete(1.0, tk.END)
    email_subject_textbox.insert(tk.END, subject1)

# E-posta gönderme fonksiyonu
def send_email():
    try:
        global email, password
        from_address = email
        password = password

        to_addresses = [email_listbox.get(i) for i in email_listbox.curselection()]
        if not to_addresses:
            messagebox.showerror("Error", "Seleccione al menos un destinatario")
            return

        if not  attached_files:
            result = messagebox.askyesno("Advertencia", "No adjuntó ningún archivo. ¿Aún desea enviar el correo electrónico?")
            if not result:
                return

        subject = email_subject_textbox.get(1.0, tk.END)
        body = email_content_textbox.get(1.0, tk.END)

        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.login(from_address, password)

        msg = MIMEMultipart()
        msg['From'] = from_address
        msg['To'] = ', '.join(to_addresses)
        msg['Subject'] = subject

        msg.attach(MIMEText(body, 'plain'))

        for file in attached_files:
            attachment = open(file, "rb")
            part = MIMEBase('application', 'octet-stream')
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(file)}")
            msg.attach(part)

        server.sendmail(from_address, to_addresses, msg.as_string())
        server.quit()
        save_to_excel()
        clear()
        messagebox.showinfo("Éxito", "Correo electrónico enviado correctamente")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# Function to save shipment information to Excel
def save_to_excel():
    try:
        excel_file = "Lista de envios.xlsx"
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook["Envios"]

        trh = datetime.now().strftime("%d/%m/%Y")
        # Add data to Excel
        sheet.append(
            [trh, palet.get(), pick_up_place.get(), place_to_delivery.get(), carrier_combobox.get(), palet_numbers()])

        workbook.save(excel_file)
        messagebox.showinfo("Éxito", "Datos guardados en Excel")
    except Exception as e:
        messagebox.showerror("Error", str(e))

############################################
# Function that loads customer data from Excel file
def load_customers():
    try:
        workbook = openpyxl.load_workbook("Lista de envios.xlsx")
        sheet = workbook["Clientes"]

        customers = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            customer_name = row[0]
            customer_tel = row[1]
            customer_address = row[2]
            customers[customer_name] = {"Telefono": customer_tel, "Direccion": customer_address}
        return customers
    except Exception as e:
        messagebox.showerror("Error", str(e))
    return {}

# # Fill in the address when selecting the customer
def fill_customer_address(event):
    customer = customer_combobox.get()
    if customer in customers:
        place_to_delivery.delete(0, tk.END)
        place_to_delivery.insert(0, customer)
        delivery_address.delete(0, tk.END)
        delivery_address.insert(0, customers[customer]["Direccion"])
        delivery_tel.delete(0, tk.END)
        delivery_tel.insert(0, customers[customer]["Telefono"])

# Fill in the address when the customer is selected in the combobox in the customer edit tab
def fill_current_address(event):
    current_customer = current_customer_combobox.get()
    if current_customer in customers:
        customer_address_entry.delete(0, tk.END)
        customer_address_entry.insert(0, current_customer)
        customer_address_entry_full.delete("1.0", tk.END)
        customer_address_entry_full.insert("1.0", customers[current_customer]["Direccion"])
        customer_tel_entry.delete(0, tk.END)
        customer_tel_entry.insert(0, customers[current_customer]["Telefono"])

# Add customer and save to Excel function
def add_customer_to_excel():
    try:
        excel_file = "Lista de envios.xlsx"
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook["Clientes"]

        customer_name = customer_address_entry.get()
        customer_tel = customer_tel_entry.get()
        customer_address = customer_address_entry_full.get(1.0, "end-1c")

        sheet.append([customer_name, customer_tel, customer_address])
        workbook.save(excel_file)

        # Add the newly added customer to the list
        customers[customer_name] = {"Telefono": customer_tel, "Direccion": customer_address}
        customer_combobox['values'] = list(customers.keys())
        customer_address_entry.delete(0, tk.END)
        customer_tel_entry.delete(0, tk.END)
        customer_address_entry_full.delete(1.0, tk.END)
        current_customer_combobox['values'] = list(customers.keys())
        messagebox.showinfo("Éxito", "Cliente añadido con éxito.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Customer information update function
def update_customer():
    excel_file = "Lista de envios.xlsx"
    workbook = openpyxl.load_workbook(excel_file)

    sheet = workbook["Clientes"]
    customer_name = current_customer_combobox.get()  # Selected customer
    new_tel = customer_tel_entry.get()
    new_address = customer_address_entry_full.get(1.0, "end-1c")

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == customer_name:
            row[1].value = new_tel
            row[2].value = new_address
            workbook.save(excel_file)  # Save changes
            customers[customer_name] = {"Telefono": new_tel, "Direccion": new_address}
            current_customer_combobox['values'] = list(customers.keys())
            messagebox.showinfo("Éxito", "¡Información del cliente actualizada!")
            break

############################################

# Adding repository and saving to Excel function
def add_warehouse_to_excel():
    try:
        excel_file = "Lista de envios.xlsx"
        workbook = openpyxl.load_workbook(excel_file)

        sheet = workbook["Almacenes"]

        warehouse_name = warehouse_name_entry.get()
        warehouse_tel = warehouse_tel_entry.get()
        warehouse_address = warehouse_address_full.get(1.0, "end-1c")

        sheet.append([warehouse_name, warehouse_tel, warehouse_address])
        workbook.save(excel_file)

        # Add the newly added customer to the list
        warehouses[warehouse_name] = {"Telefono": warehouse_tel, "Direccion": warehouse_address}
        warehouse_combobox['values'] = list(warehouses.keys())
        warehouse_name_entry.delete(0, tk.END)
        warehouse_tel_entry.delete(0, tk.END)
        warehouse_address_full.delete(1.0, tk.END)
        current_warehous_combobox['values'] = list(warehouses.keys())
        messagebox.showinfo("Éxito", "El repositorio se agregó correctamente")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Repository information update function
def update_warehouse():
    excel_file = "Lista de envios.xlsx"
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Almacenes"]

    warehouse_name = current_warehous_combobox.get()  # Selected warehouse
    new_tel = warehouse_tel_entry.get()
    new_address = warehouse_address_full.get(1.0, "end-1c")

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == warehouse_name:
            row[1].value = new_tel
            row[2].value = new_address
            workbook.save(excel_file)  # Save changes
            warehouses[warehouse_name] = {"Telefono": new_tel, "Direccion": new_address}
            current_warehous_combobox['values'] = list(warehouses.keys())
            messagebox.showinfo("Éxito", "¡Información del repositorio actualizada!")
            break

# Function that loads the Storage Addresses in the Excel file
def load_warehouses():
    try:
        workbook = openpyxl.load_workbook("Lista de envios.xlsx")
        sheet = workbook["Almacenes"]

        warehouses = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            warehouse_name = row[0]
            warehouse_tel= row[1]
            warehouse_address = row[2]
            warehouses[warehouse_name] = {"Telefono": warehouse_tel, "Direccion": warehouse_address}

        return warehouses
    except Exception as e:
        messagebox.showerror("Error", str(e))
        return {}

# Fill in the address when the warehouse is selected
def fill_warehouse_address(event):
    warehouse = warehouse_combobox.get()
    if warehouse in warehouses:
        pick_up_place.delete(0, tk.END)
        pick_up_place.insert(0, warehouse)
        pick_up_address.delete(0, tk.END)
        pick_up_address.insert(0, warehouses[warehouse]["Direccion"])
        pick_up_tel.delete(0, tk.END)
        pick_up_tel.insert(0, warehouses[warehouse]["Telefono"])

# Fill in the address when the warehouse is selected in the combobox in the warehouse editing tab
def fill_current_warehouse(event):
    current_warehous = current_warehous_combobox.get()
    if current_warehous in warehouses:
        warehouse_name_entry.delete(0, tk.END)
        warehouse_name_entry.insert(0, current_warehous)
        warehouse_address_full.delete("1.0", tk.END)
        warehouse_address_full.insert("1.0", warehouses[current_warehous]["Direccion"])
        warehouse_tel_entry.delete(0, tk.END)
        warehouse_tel_entry.insert(0, warehouses[current_warehous]["Telefono"])

############################################
# Functions for adding/editing transport mail addresses
def add_carrier():
    try:
        excel_file = "Lista de envios.xlsx"
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook["Transportistas"]

        carrier_name = new_transport_entry.get()
        carrier_mails = transport_email_entry.get(1.0, "end-1c")

        sheet.append([carrier_name, carrier_mails])
        workbook.save(excel_file)

        email_list[carrier_name] = {"Emails": carrier_mails}
        carrier_combobox['values'] = list(email_list.keys())
        mail_list_combobox['values'] = list(email_list.keys())
        new_transport_entry.delete(0, tk.END)
        transport_email_entry.delete(1.0, tk.END)
        messagebox.showinfo("Éxito", "Lista de correo electrónico agregada correctamente")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# Updates the Transport list in the mail sending tab
def update_mail_list():
    excel_file = "Lista de envios.xlsx"
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Transportistas"]
    carrier_name = mail_list_combobox.get()
    carrier_mails = transport_email_entry.get(1.0, "end-1c")

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == carrier_name:
            row[1].value = carrier_mails
            workbook.save(excel_file)  # Save changes
            email_list[carrier_name] = {"Emails": carrier_mails}
            mail_list_combobox['values'] = list(email_list.keys())
            messagebox.showinfo("Éxito", "¡Lista de correo de transporte actualizada!")
            break

# Updates the email list
def update_email_list(event):
    email_listbox.delete(0, tk.END)
    carrier_name = carrier_combobox.get()
    mail_list = [email.strip() for email in email_list[carrier_name]["Emails"].split(",")]

    if carrier_name in email_list:
        for email in mail_list:
            email_listbox.insert(tk.END, email)

# Updates the transport list in the transport add tab
def fill_current_mail_list(event):
    current_carrier = mail_list_combobox.get()
    if current_carrier in email_list:
        new_transport_entry.delete(0, tk.END)
        new_transport_entry.insert(0, current_carrier)
        transport_email_entry.delete("1.0", tk.END)
        transport_email_entry.insert("1.0", email_list[current_carrier]["Emails"])


# Function that loads the Email Addresses in the Excel file
def load_mail_list():
    try:
        workbook = openpyxl.load_workbook("Lista de envios.xlsx")
        sheet = workbook["Transportistas"]

        email_list = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            transport_name = row[0]
            emails = row[1]
            email_list[transport_name] = {"Emails": emails}

        return email_list
    except Exception as e:
        messagebox.showerror("Error", str(e))
        return {}


# Description of the text box in the add and edit transport tab
def on_text_click(event):
    """Clear placeholder when text widget is clicked."""
    if transport_email_entry.get("1.0", "end-1c") == "Ingrese direcciones de correo electrónico separadas por comas (por ejemplo: mail1@gmail.com, mail2@gmail.com)":
        transport_email_entry.delete("1.0", "end")  # Delete placeholder
        transport_email_entry.config(fg='black')  # Make text color black

def on_focusout_text(event):
    """Return placeholder if empty when exiting Text widget."""
    if transport_email_entry.get("1.0", "end-1c") == '':
        transport_email_entry.insert("1.0", "Ingrese direcciones de correo electrónico separadas por comas (por ejemplo: mail1@gmail.com, mail2@gmail.com)")
        transport_email_entry.config(fg='grey')  # Make the placeholder text gray

# File adding and removing function
attached_files = []

def clear_files():
    attached_files.clear()
    attached_files_label.config(text="")

def attach_file():
    filenames = filedialog.askopenfilenames(title="Seleccionad archivos")
    attached_files.extend(filenames)
    attached_files_label.config(text=", ".join(attached_files),anchor="w" , justify= "left", pady=10)

def clear():
    carrier_combobox.delete(0, tk.END)
    warehouse_combobox.delete(0, tk.END)
    customer_combobox.delete(0, tk.END)
    email_listbox.delete(0, tk.END)
    palet.delete(0, tk.END)
    when.delete(0, tk.END)
    pick_up_place.delete(0, tk.END)
    pick_up_address.delete(0, tk.END)
    pick_up_tel.delete(0, tk.END)
    place_to_delivery.delete(0, tk.END)
    delivery_address.delete(0, tk.END)
    delivery_tel.delete(0, tk.END)
    palet_content_textbox.delete("1.0", tk.END)
    email_subject_textbox.delete("1.0", tk.END)
    email_content_textbox.delete("1.0", tk.END)
    clear_files()


def palet_numbers():
    pedido = palet_content_textbox.get("1.0", tk.END)
    pattern = r'\b\d{5}\b|P22-\d{6}'
    results = re.findall(pattern, pedido)
    resul_filter = [item for item in results if not (len(item) == 6 and item.isdigit())]
    numbers_comma = ', '.join(resul_filter)
    return numbers_comma

# Creating an interface
root = tk.Tk()
root.title("SendTrack")
root.geometry("1000x1000")
root.minsize(850, 850)

# Adding a notebook (tab structure)
notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True)

# Pallet Shipping tab
home_frame = tk.Frame(notebook)
notebook.add(home_frame, text='Inicio')

text_frame = tk.Frame(home_frame)
text_frame.place(x=190, y=320, width=570, height=90)

# Customer Addresses tab
customer_frame = tk.Frame(notebook)
notebook.add(customer_frame, text='Clientes')

# Warehouse Addresses tab
warehouse_frame = tk.Frame(notebook)
notebook.add(warehouse_frame, text='Almacenes')

# Transport Addresses tab
transports_frame = tk.Frame(notebook)
notebook.add(transports_frame, text='Transportes')

# Settings tab
settings_frame = tk.Frame(notebook)
notebook.add(settings_frame, text="Ajustes")

########################################

# Add existing interface to Pallet Shipping Tab
# List (ComboBox) to select email recipients
tk.Label(home_frame, text="Transportista", anchor="w").place(x=10, y=10, width=170, height=20)
carrier = tk.StringVar()
carrier.set("  Jcarbo Barcelona")
carrier_combobox = ttk.Combobox(home_frame, textvariable=carrier , values=list(load_mail_list().keys()))
carrier_combobox.place(x=190, y=10, width=260, height=20)
carrier_combobox.bind("<<ComboboxSelected>>", update_email_list)

tk.Label(home_frame, text="Correos electrónicos", anchor="w").place(x=10, y=40, width=170, height=20)
email_listbox = tk.Listbox(home_frame, selectmode='multiple', height=6)
email_listbox.place(x=190, y=40, width=260, height=90)

tk.Label(home_frame, text="Cantidad de Palet", anchor="w").place(x=480, y=10, width=170, height=20)
palet = tk.Entry(home_frame)
palet.place(x=590, y=10, width=170, height=20)
tk.Button(home_frame, text="Limpiar", command=clear).place(x=770, y=10, width=150, height=20)

tk.Label(home_frame, text="Para Cuando?", anchor="w").place(x=480, y=40, width=170, height=20)
when = tk.Entry(home_frame)
when.place(x=590, y=40, width=170, height=20)

#########################################
#Warehouse
tk.Label(home_frame, text="Lugar de Recogida:", anchor="w").place(x=10, y=140, width=170, height=20)
warehouse_combobox = ttk.Combobox(home_frame, values=list(load_warehouses().keys()))
warehouse_combobox.bind("<<ComboboxSelected>>", fill_warehouse_address)
warehouse_combobox.place(x=190, y=140, width=260, height=20)

pick_up_place = tk.Entry(home_frame)
pick_up_place.place(x=590, y=200, width=170, height=20)

tk.Label(home_frame, text="Direccion de Recogida:", anchor="w").place(x=10, y=170, width=170, height=20)
pick_up_address = tk.Entry(home_frame)
pick_up_address.place(x=190, y=170, width=570, height=20)

tk.Label(home_frame, text="Telefono.:", anchor="w").place(x=480, y=140, width=170, height=20)
pick_up_tel = tk.Entry(home_frame)
pick_up_tel.place(x=590, y=140, width=170, height=20)

#########################################
#Customer
tk.Label(home_frame, text="Nombre de Cliente", anchor="w").place(x=10, y=200, width=170, height=20)
customer_combobox = ttk.Combobox(home_frame, values=list(load_customers().keys()))
customer_combobox.bind("<<ComboboxSelected>>", fill_customer_address)
customer_combobox.place(x=190, y=200, width=260, height=20)

place_to_delivery = tk.Entry(home_frame)
place_to_delivery.place(x=580, y=200, width=170, height=20)

tk.Label(home_frame, text="Direccion de Entrega", anchor="w").place(x=10, y=230, width=170, height=20)
delivery_address = tk.Entry(home_frame)
delivery_address.place(x=190, y=230, width=570, height=20)

tk.Label(home_frame, text="Numero de Telefono", anchor="w").place(x=10, y=260, width=170, height=20)
delivery_tel = tk.Entry(home_frame)
delivery_tel.place(x=190, y=260, width=570, height=20)

#########################################
# Add file button
tk.Button(home_frame, text="Agregar archivo", command=attach_file).place(x=10, y=290, width=170, height=20)
attached_files_label = tk.Label(home_frame, text="")
attached_files_label.place(x=190, y=290, width=570, height=20)

# Label for palette content
tk.Label(home_frame, text="Mercancia", anchor="w").place(x=10, y=320, width=170, height=20)

# TextBox
palet_content_textbox = tk.Text(text_frame, wrap="word")
palet_content_textbox.pack(side="left", fill="both", expand=True)

# Scrollbar
scrollbar = tk.Scrollbar(text_frame, command=palet_content_textbox.yview)
scrollbar.pack(side="right", fill="y")

# Bind Scrollbar to TextBox
palet_content_textbox.config(yscrollcommand=scrollbar.set)

# TextBox for email subject
tk.Label(home_frame, text="Asunto", anchor="w").place(x=10, y=420, width=170, height=20)
email_subject_textbox = tk.Text(home_frame, height=20, width=80)
email_subject_textbox.place(x=190, y=420, width=650, height=20)

# TextBox for email content
tk.Label(home_frame, text="Mensaje", anchor="w").place(x=10, y=450, width=170, height=20)
email_content_textbox = tk.Text(home_frame, height=20, width=80)
email_content_textbox.place(x=190, y=450, width=650, height=300)

# Button that updates the email content
tk.Button(home_frame, text="Actualizar el contenido", command=update_email_content).place(x=10, y=450, width=170, height=20)
# Send email button
tk.Button(home_frame, text="Enviar", command=send_email).place(x=10, y=480, width=170, height=20)
# Remove selected file button
tk.Button(home_frame, text="Eliminar archivo seleccionado", command=clear_files).place(x=10, y=510, width=170, height=20)

########################################
# Customer Addresses editing field
tk.Label(customer_frame, text="Selecciona el cliente", anchor="w").place(x=10, y=10, width=170, height=20)
current_customer_combobox = ttk.Combobox(customer_frame, values=list(load_customers().keys()))
current_customer_combobox.bind("<<ComboboxSelected>>", fill_current_address)
current_customer_combobox.place(x=150, y=10, width=260, height=20)

tk.Label(customer_frame, text="Nuevo Cliente", anchor="w").place(x=10, y=40, width=170, height=20)
customer_address_entry= tk.Entry(customer_frame)
customer_address_entry.place(x=150, y=40, width=260, height=20)

tk.Label(customer_frame, text="Numero de Telefono", anchor="w").place(x=10, y=70, width=170, height=20)
customer_tel_entry = tk.Entry(customer_frame)
customer_tel_entry.place(x=150, y=70, width=260, height=20)

tk.Label(customer_frame, text="Direccion", anchor="w").place(x=10, y=100, width=170, height=20)
customer_address_entry_full = tk.Text(customer_frame, wrap="word", padx=5, pady=5)
customer_address_entry_full.place(x=150, y=100, width=260, height=130)

tk.Button(customer_frame, text="Añadir Cliente", command=add_customer_to_excel, anchor="w").place(x=150, y=240, width=120, height=20)
tk.Button(customer_frame, text="Actualizar el Cliente", command=update_customer, anchor="w").place(x=280, y=240, width=120, height=20)

###############################################
# Storage Addresses editing field

tk.Label(warehouse_frame, text="Nombre de Almacen:", anchor="w").place(x=10, y=10, width=170, height=20)
current_warehous_combobox = ttk.Combobox(warehouse_frame, values=list(load_warehouses().keys()))
current_warehous_combobox.bind("<<ComboboxSelected>>", fill_current_warehouse)
current_warehous_combobox.place(x=150, y=10, width=260, height=20)

tk.Label(warehouse_frame, text="Nuevo Almacen", anchor="w").place(x=10, y=40, width=170, height=20)
warehouse_name_entry = tk.Entry(warehouse_frame)
warehouse_name_entry.place(x=150, y=40, width=260, height=20)

tk.Label(warehouse_frame, text="Telefono", anchor="w").place(x=10, y=70, width=170, height=20)
warehouse_tel_entry = tk.Entry(warehouse_frame)
warehouse_tel_entry.place(x=150, y=70, width=260, height=20)

tk.Label(warehouse_frame, text="Direccion", anchor="w").place(x=10, y=100, width=170, height=20)
warehouse_address_full = tk.Text(warehouse_frame, wrap="word", padx=5, pady=5)
warehouse_address_full.place(x=150, y=100, width=260, height=130)

tk.Button(warehouse_frame, text="Añadir Nave", command=add_warehouse_to_excel, anchor="w").place(x=150, y=240, width=120, height=20)
tk.Button(warehouse_frame, text="Actualizar Nave", command=update_warehouse, anchor="w").place(x=280, y=240, width=120, height=20)

########################################

# Transport Addresses editing area
tk.Label(transports_frame, text="Nombre de Transportista:", anchor="w").place(x=10, y=10, width=170, height=20)
mail_list_combobox = ttk.Combobox(transports_frame, values=list(load_mail_list().keys()))
mail_list_combobox.bind("<<ComboboxSelected>>", fill_current_mail_list)
mail_list_combobox.place(x=150, y=10, width=260, height=20)

tk.Label(transports_frame, text="Nuevo Transportista:", anchor="w").place(x=10, y=40, width=170, height=20)
new_transport_entry = tk.Entry(transports_frame)
new_transport_entry.place(x=150, y=40, width=260, height=20)

tk.Label(transports_frame, text="Correos del Transportista:", anchor="w").place(x=10, y=70, width=170, height=20)
transport_email_entry = tk.Text(transports_frame, wrap="word", padx=5, pady=5)
transport_email_entry.place(x=150, y=70, width=350, height=130)
transport_email_entry.insert(1.0, "Ingrese direcciones de correo electrónico separadas por comas (por ejemplo: mail1@gmail.com, mail2@gmail.com)")

# Delete placeholder and allow writing when entry is clicked
transport_email_entry.bind('<FocusIn>', on_text_click)
# When exiting the entry, return the placeholder if it is empty.
transport_email_entry.bind('<FocusOut>', on_focusout_text)

tk.Button(transports_frame, text="Añadir Transportista", command=add_carrier, anchor="w").place(x=150, y=210, width=120, height=20)
tk.Button(transports_frame, text="Actualizar", command=update_mail_list, anchor="w").place(x=280, y=210, width=120, height=20)

# Settings area
# Enter your email address
tk.Label(settings_frame, text="Dirección de Correo Electrónico de Remitente:").pack(pady=5)
from_address_entry = tk.Entry(settings_frame, width=40)
from_address_entry.pack(pady=5)

# Password entry
tk.Label(settings_frame, text="Contreseña:").pack(pady=5)
password_entry = tk.Entry(settings_frame, show="*", width=40)
password_entry.pack(pady=5)

# Save button
save_button = tk.Button(settings_frame, text="Guardar", command=save_credentials)
save_button.pack(pady=20)

# Load data when application starts
customers = load_customers()
customer_combobox['values'] = sorted(customers.keys())
current_customer_combobox['values'] = sorted(customers.keys())

warehouses = load_warehouses()
warehouse_combobox['values'] = sorted(warehouses.keys())
current_warehous_combobox['values'] = sorted(warehouses.keys())

email_list = load_mail_list()
carrier_combobox['values'] = sorted(email_list.keys())
mail_list_combobox['values'] = sorted(email_list.keys())

# Load email and password when app starts
load_credentials()

root.mainloop()